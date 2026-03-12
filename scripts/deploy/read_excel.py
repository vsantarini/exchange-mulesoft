import openpyxl
import json
import argparse

SHEET_DEPLOYMENTS = "Deployments"
SHEET_POLICIES = "Policies"

def read_sheet(ws):
 headers = [cell.value for cell in ws[1]]
 rows = []
 for row in ws.iter_rows(min_row=2, values_only=True):
 record = dict(zip(headers, row))
 if any(record.values()):
 rows.append(record)
 return rows

if __name__ == "__main__":
 parser = argparse.ArgumentParser()
 parser.add_argument("--file", required=True)
 parser.add_argument("--output-deployments", required=True)
 parser.add_argument("--output-policies", required=True)
 args = parser.parse_args()

 wb = openpyxl.load_workbook(args.file)

 deployments = read_sheet(wb[SHEET_DEPLOYMENTS])
 policies = read_sheet(wb[SHEET_POLICIES])

 with open(args.output_deployments, "w") as f:
 json.dump(deployments, f, indent=2)
 with open(args.output_policies, "w") as f:
 json.dump(policies, f, indent=2)

 print(f"✅ Deployments: {len(deployments)} | Policies: {len(policies)}")