import openpyxl
import json
import argparse

def read_excel(file_path):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    apis = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        api = dict(zip(headers, row))
        if api.get("assetId"):
            apis.append(api)
    return apis

# Expected Excel columns:
# assetId | name | version | apiVersion | type (rest/soap) |
# filePath | imageFile | tags | categories | description

if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--file",   required=True)
    parser.add_argument("--output", required=True)
    args = parser.parse_args()

    apis = read_excel(args.file)
    with open(args.output, "w") as f:
        json.dump(apis, f, indent=2)
    print(f"✅ Found {len(apis)} APIs in Excel.")