import openpyxl
import json
import argparse
import os
import sys

SHEET_APIS = "APIs"
SHEET_APPS = "Applications"
SHEET_CERTS = "Certificates"

def read_sheet(ws):
    """Legge un foglio Excel e restituisce una lista di dizionari (un record per riga)."""
    headers = [cell.value for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        record = dict(zip(headers, row))
        if any(v is not None for v in record.values()):
            rows.append(record)
    return rows

def normalize_cert_record(record):
    """Normalizza e valida un record del foglio Certificates."""
    VALID_FORMATS = ["PEM", "CRT", "DER", "P12", "PFX", "JKS"]
    VALID_ENVS = ["dev", "test", "uat", "prod"]

    # ── Normalizzazione campi stringa ──
    for key in ["appName", "certAlias", "certFilePath", "certSubjectDN",
                "secretGroupName", "truststoreName", "notes"]:
        record[key] = str(record.get(key) or "").strip()

    # ── Formato certificato ──
    fmt = str(record.get("certFormat") or "").strip().upper()
    if fmt not in VALID_FORMATS:
        print(f"[WARN] Formato '{fmt}' non valido per alias '{record['certAlias']}'. "
              f"Valori ammessi: {VALID_FORMATS}")
    record["certFormat"] = fmt

    # ── Password (può essere vuota per PEM/CRT/DER) ──
    record["certPassword"] = str(record.get("certPassword") or "").strip()

    # ── Ambiente target ──
    env = str(record.get("targetEnv") or "").strip().lower()
    if env not in VALID_ENVS:
        print(f"[WARN] Ambiente '{env}' non valido per alias '{record['certAlias']}'. "
              f"Valori ammessi: {VALID_ENVS}")
    record["targetEnv"] = env

    # ── Flag useAsClientId ──
    use_flag = str(record.get("useAsClientId") or "FALSE").strip().upper()
    record["useAsClientId"] = use_flag == "TRUE"

    # ── Data di scadenza ──
    exp = record.get("expirationDate")
    if exp is not None and hasattr(exp, "isoformat"):
        record["expirationDate"] = exp.isoformat()
    else:
        record["expirationDate"] = str(exp or "")

    return record

def enrich_apps_with_certs(apps, certs):
    """Arricchisce i record delle Applications con i certificati associati per appName."""
    cert_map = {}
    for cert in certs:
        app_name = cert.get("appName", "")
        if app_name:
            cert_map.setdefault(app_name, []).append(cert)

    for app in apps:
        name = app.get("appName", "")
        app["certificates"] = cert_map.get(name, [])

    return apps

if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Legge il catalogo API/App/Certificati dall'Excel di input"
    )
    parser.add_argument("--file", required=True,
                        help="Path al file Excel di input (api-catalog.xlsx)")
    parser.add_argument("--output-apis", required=True,
                        help="File JSON di output per le API")
    parser.add_argument("--output-apps", required=True,
                        help="File JSON di output per le Applications")
    parser.add_argument("--output-certs", required=False, default="cert-list.json",
                        help="File JSON di output per i Certificati (default: cert-list.json)")
    args = parser.parse_args()

    # ── Validazione file di input ──
    if not os.path.isfile(args.file):
        print(f"[ERROR] File non trovato: {args.file}")
        sys.exit(1)

    wb = openpyxl.load_workbook(args.file)

    # ── Lettura foglio APIs ──
    if SHEET_APIS not in wb.sheetnames:
        print(f"[ERROR] Foglio '{SHEET_APIS}' non trovato nel file Excel.")
        sys.exit(1)
    apis = read_sheet(wb[SHEET_APIS])

    # ── Lettura foglio Applications ──
    if SHEET_APPS not in wb.sheetnames:
        print(f"[ERROR] Foglio '{SHEET_APPS}' non trovato nel file Excel.")
        sys.exit(1)
    apps = read_sheet(wb[SHEET_APPS])

    # ── Lettura foglio Certificates (opzionale per retrocompatibilità) ──
    certs = []
    if SHEET_CERTS in wb.sheetnames:
        raw_certs = read_sheet(wb[SHEET_CERTS])
        certs = [normalize_cert_record(c) for c in raw_certs]
        print(f"[OK] Certificates: {len(certs)} record letti")

        # Arricchisci le Applications con i certificati associati
        apps = enrich_apps_with_certs(apps, certs)
    else:
        print(f"[WARN] Foglio '{SHEET_CERTS}' non presente. "
              f"Nessun certificato verrà processato.")

    # ── Salvataggio output ──
    with open(args.output_apis, "w") as f:
        json.dump(apis, f, indent=2)

    with open(args.output_apps, "w") as f:
        json.dump(apps, f, indent=2)

    with open(args.output_certs, "w") as f:
        json.dump(certs, f, indent=2)

    print(f"[OK] APIs: {len(apis)} | Apps: {len(apps)} | Certificates: {len(certs)}")